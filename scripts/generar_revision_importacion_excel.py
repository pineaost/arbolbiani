"""Genera una propuesta conservadora de altas desde el Excel genealógico.

No accede a Supabase ni crea relaciones. El resultado sirve para revisar la
interpretación antes de ejecutar cualquier inserción.
"""

from __future__ import annotations

import json
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


RAIZ = Path(__file__).resolve().parents[1]
SALIDA = RAIZ / "Referencias" / "importacion-arbol-genealogico-revision.json"
COLUMNAS = {
    "nombre": 1,
    "nacimiento": 2,
    "lugar_nacimiento": 3,
    "defuncion": 4,
    "lugar_defuncion": 5,
    "profesion": 13,
    "direcciones": 14,
    "notas": 15,
    "informacion": 16,
}
SUFIJOS_APELLIDO = ("della paolera", "della paoelra", "di gregorio", "de luque")


def texto(valor: Any) -> str | None:
    if valor is None:
        return None
    resultado = str(valor).strip()
    return resultado or None


def normalizar_clave(valor: str) -> str:
    import unicodedata

    sin_tildes = "".join(
        caracter for caracter in unicodedata.normalize("NFD", valor)
        if unicodedata.category(caracter) != "Mn"
    )
    return re.sub(r"\s+", " ", sin_tildes).strip().casefold()


def titulo(valor: str) -> str:
    return " ".join(parte[:1].upper() + parte[1:].lower() for parte in valor.split())


def separar_nombre(nombre_completo: str) -> tuple[str, str]:
    limpio = re.sub(r"\s+", " ", nombre_completo).strip()
    minusculas = limpio.casefold()
    for sufijo in SUFIJOS_APELLIDO:
        if minusculas.endswith(sufijo):
            inicio = len(limpio) - len(sufijo)
            return titulo(limpio[:inicio].strip()), titulo(limpio[inicio:].strip())
    partes = limpio.split(" ")
    if len(partes) < 2:
        raise ValueError(f"No se puede separar nombre y apellido: {nombre_completo!r}")
    return titulo(" ".join(partes[:-1])), titulo(partes[-1])


def fecha_precisa(valor: Any) -> tuple[str | None, str | None]:
    """Devuelve ISO sólo si la fuente da día, mes y año inequívocos."""
    if isinstance(valor, datetime):
        return valor.date().isoformat(), None
    if isinstance(valor, date):
        return valor.isoformat(), None
    original = texto(valor)
    if not original:
        return None, None
    coincidencia = re.fullmatch(r"(\d{1,2})/(\d{1,2})/(\d{4})", original)
    if coincidencia:
        dia, mes, anio = (int(parte) for parte in coincidencia.groups())
        try:
            return date(anio, mes, dia).isoformat(), None
        except ValueError:
            pass
    return None, original


def fuente(hoja: str, fila: int, columna: str) -> dict[str, Any]:
    return {"hoja": hoja, "fila": fila, "columna": columna}


def notas_primarias(fila: tuple[Any, ...], fecha_nacimiento_no_precisa: str | None, fecha_defuncion_no_precisa: str | None) -> str | None:
    partes: list[str] = []
    if fecha_nacimiento_no_precisa:
        partes.append(f"Fecha de nacimiento consignada sin precisión suficiente: {fecha_nacimiento_no_precisa}.")
    if fecha_defuncion_no_precisa:
        partes.append(f"Fecha de defunción consignada sin precisión suficiente: {fecha_defuncion_no_precisa}.")
    for etiqueta, indice in (
        ("Profesión", COLUMNAS["profesion"]),
        ("Direcciones registradas", COLUMNAS["direcciones"]),
        ("Notas del Excel", COLUMNAS["notas"]),
        ("Información importante", COLUMNAS["informacion"]),
    ):
        valor = texto(fila[indice] if indice < len(fila) else None)
        if valor:
            partes.append(f"{etiqueta}: {valor}")
    return "\n\n".join(partes) or None


def crear_persona(
    nombre_completo: str,
    fuentes: list[dict[str, Any]],
    *,
    fecha_nacimiento: Any = None,
    lugar_nacimiento: str | None = None,
    fecha_fallecimiento: Any = None,
    lugar_fallecimiento: str | None = None,
    observaciones: list[str] | None = None,
) -> dict[str, Any]:
    nombre, apellido = separar_nombre(nombre_completo)
    nacimiento, nacimiento_no_precisa = fecha_precisa(fecha_nacimiento)
    fallecimiento, fallecimiento_no_precisa = fecha_precisa(fecha_fallecimiento)
    notas = list(observaciones or [])
    if nacimiento_no_precisa:
        notas.append(f"Fecha de nacimiento consignada sin precisión suficiente: {nacimiento_no_precisa}.")
    if fallecimiento_no_precisa:
        notas.append(f"Fecha de defunción consignada sin precisión suficiente: {fallecimiento_no_precisa}.")
    return {
        "clave": normalizar_clave(f"{nombre} {apellido}"),
        "persona": {
            "nombre": nombre,
            "apellido": apellido,
            "genero": "no_definido",
            "fecha_nacimiento": nacimiento,
            "lugar_nacimiento": lugar_nacimiento,
            "fecha_fallecimiento": fallecimiento,
            "lugar_fallecimiento": lugar_fallecimiento,
            "notas": "\n\n".join(notas) or None,
        },
        "fuentes": fuentes,
    }


def leer_principales(libro: Any) -> list[dict[str, Any]]:
    personas: list[dict[str, Any]] = []
    for hoja in libro.worksheets:
        for numero_fila, fila in enumerate(hoja.iter_rows(values_only=True), start=1):
            nombre_completo = texto(fila[COLUMNAS["nombre"]] if len(fila) > COLUMNAS["nombre"] else None)
            if numero_fila == 1 or not nombre_completo:
                continue
            nacimiento, nacimiento_no_precisa = fecha_precisa(fila[COLUMNAS["nacimiento"]])
            fallecimiento, fallecimiento_no_precisa = fecha_precisa(fila[COLUMNAS["defuncion"]])
            nombre, apellido = separar_nombre(nombre_completo)
            personas.append({
                "clave": normalizar_clave(f"{nombre} {apellido}"),
                "persona": {
                    "nombre": nombre,
                    "apellido": apellido,
                    "genero": "no_definido",
                    "fecha_nacimiento": nacimiento,
                    "lugar_nacimiento": texto(fila[COLUMNAS["lugar_nacimiento"]]),
                    "fecha_fallecimiento": fallecimiento,
                    "lugar_fallecimiento": texto(fila[COLUMNAS["lugar_defuncion"]]),
                    "notas": notas_primarias(fila, nacimiento_no_precisa, fallecimiento_no_precisa),
                },
                "fuentes": [fuente(hoja.title, numero_fila, "Nombre")],
            })
    return personas


# Estas menciones fueron seleccionadas después de revisar las columnas
# relacionales. Sólo incluyen personas con nombre y apellido explícitos.
DERIVADAS = [
    ("Domenica Bertola", [fuente("Biani", 5, "Nombre de la madre")], {}),
    ("Rosa Ferri", [fuente("Biani", 5, "Nombre del cónyuge"), fuente("Biani", 6, "Nombre de la madre")], {"observaciones": ["También aparece como “Rosa Serafina” en la fuente."]}),
    ("María Biasotti", [fuente("Biani", 6, "Nombre del cónyuge")], {}),
    ("María Magdalena Milani", [fuente("Biani", 7, "Nombre del cónyuge"), fuente("Biani", 8, "Nombre de la madre")], {"fecha_nacimiento": "3/3/1893", "lugar_nacimiento": "Alcorta", "fecha_fallecimiento": "25/2/1970", "lugar_fallecimiento": "Victoria"}),
    ("Gustavo Ricardo Biani", [fuente("Biani", 8, "Nombres de los hijos"), fuente("Podrecca", 8, "Nombres de los hijos")], {"fecha_nacimiento": "8/5/1959"}),
    ("Graciela Bernardita Biani", [fuente("Biani", 8, "Nombres de los hijos"), fuente("Podrecca", 8, "Nombres de los hijos")], {"fecha_nacimiento": "17/5/1961?"}),
    ("Josefa Varela", [fuente("Acevey", 2, "Nombre del cónyuge"), fuente("Acevey", 3, "Nombre de la madre")], {"fecha_nacimiento": "1740-xxxx"}),
    ("Francisca Ávalos", [fuente("Acevey", 3, "Nombre del cónyuge"), fuente("Acevey", 4, "Nombre de la madre")], {}),
    ("Martina Correa", [fuente("Acevey", 4, "Nombre del cónyuge"), fuente("Acevey", 5, "Nombre de la madre")], {"fecha_nacimiento": "1780-xxxx"}),
    ("Josefa Burgos", [fuente("Acevey", 5, "Nombre del cónyuge"), fuente("Acevey", 6, "Nombre de la madre")], {"fecha_nacimiento": "1815-xxxx"}),
    ("Emilia Espinosa", [fuente("Acevey", 6, "Nombre del cónyuge"), fuente("Acevey", 7, "Nombre de la madre")], {"fecha_nacimiento": "14/8/1840"}),
    ("Matilde Florentina Lopez", [fuente("Acevey", 7, "Nombre del cónyuge"), fuente("Acevey", 8, "Nombre de la madre")], {}),
    ("Nelly Cerliani", [fuente("Acevey", 8, "Nombre del cónyuge")], {}),
    ("Irene Angélica della Paolera", [fuente("Acevey", 8, "Nombre del cónyuge")], {}),
    ("Susana Marta Acevey", [fuente("Acevey", 8, "Nombres de los hijos")], {}),
    ("Graciela Mirta Acevey", [fuente("Acevey", 8, "Nombres de los hijos")], {}),
    ("Vicenza Buonotempo", [fuente("Della Paolera", 3, "Nombre del cónyuge"), fuente("Della Paolera", 4, "Nombre de la madre")], {"fecha_nacimiento": "1764-xxxx"}),
    ("Mariantonieta di Gregorio", [fuente("Della Paolera", 4, "Nombre del cónyuge"), fuente("Della Paolera", 5, "Nombre de la madre")], {}),
    ("Catalina Costa", [fuente("Della Paolera", 5, "Nombre del cónyuge"), fuente("Della Paolera", 6, "Nombre de la madre")], {"fecha_nacimiento": "1862-xxxx"}),
    ("José Costa", [fuente("Della Paolera", 5, "Nombre del cónyuge")], {}),
    ("Clara Agustino", [fuente("Della Paolera", 5, "Nombre del cónyuge")], {}),
    ("Luisa Soriano", [fuente("Della Paolera", 6, "Nombre del cónyuge")], {"fecha_fallecimiento": "6/7/1924", "observaciones": ["Fecha de nacimiento consignada sin estructurar: 5/10/1989. Resulta cronológicamente incompatible con el resto de la fila y no se corrige por inferencia."]}),
    ("María Francisca Ezequiela de Luque", [fuente("Della Paolera", 6, "Nombre del cónyuge")], {"fecha_nacimiento": "1897-xxxx"}),
    ("Peregrina Catalina Lidia della Paolera", [fuente("Della Paolera", 5, "Nombres de los hijos"), fuente("Della Paolera", 6, "Nombre de los hermanos")], {"fecha_nacimiento": "1893-1944"}),
    ("Teresa Ruibal", [fuente("Della Paolera", 6, "Nombre del cónyuge")], {}),
    ("Antonio Soriano", [fuente("Della Paolera", 6, "Nombre del cónyuge")], {}),
    ("Catterina Venturini", [fuente("Podrecca", 5, "Nombre del cónyuge"), fuente("Podrecca", 6, "Nombre de la madre")], {"observaciones": ["La fuente indica que murió antes de 1898; no se crea una fecha exacta."]}),
    ("Marianna Birtig", [fuente("Podrecca", 6, "Nombre del cónyuge"), fuente("Podrecca", 7, "Nombre de la madre")], {"fecha_nacimiento": "1875-xxxx"}),
    ("Livia Ermenegilda Carniello", [fuente("Podrecca", 7, "Nombre del cónyuge"), fuente("Podrecca", 8, "Nombre de la madre")], {}),
    ("Anna Maria Podrecca", [fuente("Podrecca", 6, "Nombres de los hijos")], {"fecha_nacimiento": "23/10/1898", "fecha_fallecimiento": "26/4/1899"}),
    ("Bruno Podrecca", [fuente("Podrecca", 8, "Nombre de los hermanos")], {}),
]

EXISTENTES = {
    "francesco biani": {"id": "reconocido-en-archivo", "motivo": "Nombre y año de nacimiento coinciden con Archivo Familiar."},
    "giovanni biani": {"id": "reconocido-en-archivo", "motivo": "Nombre, fecha y lugar de nacimiento coinciden con Archivo Familiar."},
}


def main() -> None:
    archivos = list(RAIZ.glob("*.xlsx"))
    if len(archivos) != 1:
        raise RuntimeError(f"Se esperaba un único Excel en la raíz; se encontraron: {archivos}")
    libro = load_workbook(archivos[0], data_only=True, read_only=True)
    personas = leer_principales(libro)
    personas.extend(crear_persona(nombre, fuentes, **datos) for nombre, fuentes, datos in DERIVADAS)

    if len(personas) != 52:
        raise RuntimeError(f"La revisión esperaba 52 personas candidatas y obtuvo {len(personas)}.")

    for indice, registro in enumerate(personas, start=1):
        registro["id_propuesta"] = f"excel-{indice:03d}"
        existente = EXISTENTES.get(registro["clave"])
        registro["estado"] = "existente_reconocido" if existente else "propuesta_importar"
        if existente:
            registro["coincidencia_en_archivo"] = existente
        del registro["clave"]

    salida = {
        "fuente": {"archivo": archivos[0].name, "hojas_revisadas": [hoja.title for hoja in libro.worksheets]},
        "criterio": {
            "relaciones": "No se importan ni se infieren relaciones de filiación, cónyuge o hermanos.",
            "fechas": "Sólo se cargan en campos de fecha los valores con día, mes y año inequívocos. Años, aproximaciones y dudas permanecen en notas.",
            "nombres_relacionales": "Sólo se proponen menciones con nombre y apellido explícitos; las listas que sólo contienen nombres de pila quedan para revisión manual.",
            "genero": "Se conserva no_definido porque el Excel no provee un campo de género confiable.",
        },
        "resumen": {
            "filas_principales_con_nombre": 21,
            "personas_candidatas_detectadas": len(personas),
            "existentes_reconocidos_sin_modificar": sum(persona["estado"] == "existente_reconocido" for persona in personas),
            "propuestas_de_alta": sum(persona["estado"] == "propuesta_importar" for persona in personas),
        },
        "personas": personas,
        "posibles_duplicados": [
            {"mencion": "Juan Luis Acevey", "fuente": fuente("Acevey", 5, "Nombre del padre"), "posible_coincidencia": "Juan Luis Acebey (Acevey/Acebey)", "decision": "No crear una ficha adicional; requiere revisión ortográfica antes de unificar."},
            {"mencion": "Michellangelo della Polera", "fuente": fuente("Della Paolera", 4, "Nombre del padre"), "posible_coincidencia": "Michellangelo Della Paoelra", "decision": "Se trata como variante de escritura de la fila principal; no se crea otra ficha."},
            {"mencion": "Próspero della Paolera", "fuente": fuente("Della Paolera", 5, "Nombre del padre"), "posible_coincidencia": "Próspero Della Paolera", "decision": "Se trata como variante de mayúsculas y acento de la fila principal; no se crea otra ficha."},
            {"mencion": "Geremia Giuseppe Podreca", "fuente": fuente("Podrecca", 8, "Nombre del padre"), "posible_coincidencia": "Geremia Giuseppe Podrecca", "decision": "Se trata como variante de una letra de la fila principal; no se crea otra ficha."},
            {"mencion": "Giuseppe Podrecca", "fuente": fuente("Podrecca", 7, "Nombre del padre"), "posible_coincidencia": "Giovanni Podrecca (dos filas principales)", "decision": "Ambiguo: no se crea ni se fusiona una ficha."},
        ],
        "datos_ambiguos_u_omitidos": [
            {"fuentes": [fuente("Biani", 5, "Nombres de los hijos"), fuente("Biani", 6, "Nombre de los hermanos"), fuente("Biani", 6, "Nombres de los hijos"), fuente("Biani", 7, "Nombre de los hermanos"), fuente("Biani", 7, "Nombres de los hijos")], "motivo": "Listas con nombres de pila sin apellido (por ejemplo Antonia, Paolo, José Valentín). No se infiere un apellido."},
            {"fuentes": [fuente("Acevey", 3, "Nombres de los hijos"), fuente("Acevey", 4, "Nombre de los hermanos"), fuente("Acevey", 4, "Nombres de los hijos"), fuente("Acevey", 5, "Nombres de los hijos"), fuente("Acevey", 6, "Nombres de los hijos"), fuente("Acevey", 7, "Nombre de los hermanos"), fuente("Acevey", 7, "Nombres de los hijos"), fuente("Acevey", 8, "Nombre de los hermanos")], "motivo": "Listas con nombres de pila, abreviaturas (NN) o apellidos no explícitos. Sólo se seleccionaron los nombres completos de Susana y Graciela Acevey."},
            {"fuentes": [fuente("Della Paolera", 3, "Nombres de los hijos"), fuente("Della Paolera", 4, "Nombre de los hermanos"), fuente("Della Paolera", 4, "Nombres de los hijos"), fuente("Della Paolera", 5, "Nombre de los hermanos"), fuente("Della Paolera", 5, "Nombres de los hijos"), fuente("Della Paolera", 6, "Nombres de los hijos")], "motivo": "Listas con nombres de pila sin apellido. Sólo se propuso Peregrina Catalina Lidia della Paolera porque su apellido sí está explícito."},
            {"fuentes": [fuente("Podrecca", 6, "Nombres de los hijos"), fuente("Podrecca", 8, "Nombres de los hijos")], "motivo": "Se dejó sin alta María Teresa Catalina e Irene Angélica de la lista de hijos porque no traen apellido; Irene Angélica della Paolera se propone únicamente por la mención completa de Acevey."},
            {"fuentes": [fuente("Della Paolera", 6, "Nombre del cónyuge")], "motivo": "La fecha “5/10/1989” de Luisa Soriano resulta inconsistente con el contexto cronológico. Se conserva como nota, no como fecha estructurada."},
        ],
    }
    SALIDA.write_text(json.dumps(salida, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(salida["resumen"], ensure_ascii=False))


if __name__ == "__main__":
    main()
